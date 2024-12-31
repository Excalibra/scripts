import json
import time
import requests
import os
from datetime import datetime
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from snownlp import SnowNLP
import statistics
import jieba
from wordcloud import WordCloud
import matplotlib.pyplot as plt
import platform
import thulac
import matplotlib.font_manager as fm
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By


'''''''''

# Reference Links

## General

Regex: https://regex101.com/
Zhihu - Two ways to obtain Bilibili video bullet comments using Python: https://zhuanlan.zhihu.com/p/609154366
Juejin - Parsing Bilibili video bullet comments: https://juejin.cn/post/7137928570080329741
CSDN - Bilibili historical bullet comment crawler: https://blog.csdn.net/sinat_18665801/article/details/104519838
CSDN - How to write a Bilibili bullet comment crawler: https://blog.csdn.net/bigbigsman/article/details/78639053?utm_source=app
Bilibili - Bilibili bullet comment notes: https://www.bilibili.com/read/cv5187469/
Bilibili third-party API: https://www.bookstack.cn/read/BilibiliAPIDocs/README.md

## Reverse Lookup by UID

https://github.com/esterTion/BiliBili_crc2mid
https://github.com/cwuom/GetDanmuSender/blob/main/main.py
https://github.com/Aruelius/crc32-crack

## User Basic Information

https://api.bilibili.com/x/space/acc/info?mid=298220126
https://github.com/ria-klee/bilibili-uid
https://github.com/SocialSisterYi/bilibili-API-collect/blob/master/docs/user/space.md

## Comments

https://www.bilibili.com/read/cv10120255/
https://github.com/SocialSisterYi/bilibili-API-collect/blob/master/docs/comment/readme.md

## JSON

https://json-schema.apifox.cn
https://bbs.huaweicloud.com/blogs/279515
https://www.cnblogs.com/mashukui/p/16972826.html

## Cookie

https://developer.mozilla.org/zh-CN/docs/Web/HTTP/Cookies

## Unpacking

https://www.cnblogs.com/will-wu/p/13251545.html
https://www.w3schools.com/python/python_tuples.asp

'''''''''''

class BilibiliAPI:
    @staticmethod
    # Parse video link basic information JSON and return it in JSON format
    def get_bv_json(video_url):
        video_id = re.findall(r'BV\w+', video_url)[0]
        api_url = f'https://api.bilibili.com/x/web-interface/view?bvid={video_id}'
        bv_json = requests.get(api_url).json()
        return bv_json

    @staticmethod
    # Parse video link bullet comments XML using the 'cid' field in JSON
    def get_danmu_xml(bv_json):
        cid = bv_json['data']["cid"]
        api_url = f'https://comment.bilibili.com/{cid}.xml'
        danmu_xml = api_url
        return danmu_xml

    @staticmethod
    # Parse video link comments JSON using the 'aid' field in JSON
    def get_comment_json(bv_json):
        aid = bv_json['data']["aid"]
        api_url = f'https://api.bilibili.com/x/v2/reply/main?next=1&type=1&oid={aid}'
        comment_json = requests.get(api_url).json()
        return comment_json

    @staticmethod
    # Enhanced parsing of video link comments JSON using the 'aid' field in JSON
    def get_comment_json_to_webui(bv_json):
        aid = bv_json['data']["aid"]
        api_url = f'https://api.bilibili.com/x/v2/reply/main?next=1&type=1&oid={aid}'

        # Determine the current operating system type
        if platform.system() == "Windows":
            # Windows platform
            driver = webdriver.Chrome()
        else:
            # Other platforms
            driver = webdriver.Chrome(ChromeDriverManager().install())

        # Provide login time
        print("Provide 45 seconds for Bilibili login")
        time.sleep(45)

        # Open the link
        driver.get(api_url)

        # Provide view effect time
        print("Provide 15 seconds to check the effects")
        time.sleep(15)

        # Find the <pre> element
        pre_element = driver.find_element(By.TAG_NAME, 'pre')

        # Get the text content of the element
        text_content = pre_element.text

        # Close WebDriver
        driver.quit()

        return text_content

    @staticmethod
    # Traverse user information and return basic parameters, preparing for XLSX write-in
    def get_user_card(mid, cookies):
            api_url = f'https://account.bilibili.com/api/member/getCardByMid?mid={mid}'
            try:
                response = requests.get(api_url, cookies=cookies)
                user_card_json = response.json()
            except json.JSONDecodeError:
                return {"error": "Failed to parse JSON. Ensure a good network environment. Too many API calls might trigger restrictions; try again later."}

            if 'message' in user_card_json:
                message = user_card_json['message']
                if 'request blocked' in message or 'frequent requests' in message:
                    return {"warning": "Ensure a good network environment. Too many API calls might trigger restrictions; try again later."}

            return user_card_json

class CRC32Checker:
    ''''''''''
    # CRC32 cracking
    # Source: https://github.com/Aruelius/crc32-crack
    # Author: Aruelius
    # Note: This section has been slightly adjusted and encapsulated as a class for easier use.
    '''''''''

    CRCPOLYNOMIAL = 0xEDB88320
    crctable = [0 for x in range(256)]

    def __init__(self):
        self.create_table()

    def create_table(self):
        # Create a CRC table for quick CRC value computation
        for i in range(256):
            crcreg = i
            for _ in range(8):
                if (crcreg & 1) != 0:
                    crcreg = self.CRCPOLYNOMIAL ^ (crcreg >> 1)
                else:
                    crcreg = crcreg >> 1
            self.crctable[i] = crcreg

    def crc32(self, string):
        # Compute the CRC32 value for the given string
        crcstart = 0xFFFFFFFF
        for i in range(len(str(string))):
            index = (crcstart ^ ord(str(string)[i])) & 255
            crcstart = (crcstart >> 8) ^ self.crctable[index]
        return crcstart

    def crc32_last_index(self, string):
        # Compute the last character CRC table index for a given string
        crcstart = 0xFFFFFFFF
        for i in range(len(str(string))):
            index = (crcstart ^ ord(str(string)[i])) & 255
            crcstart = (crcstart >> 8) ^ self.crctable[index]
        return index

    def get_crc_index(self, t):
        # Find the index in the CRC table corresponding to the highest byte value
        for i in range(256):
            if self.crctable[i] >> 24 == t:
                return i
        return -1

    def deep_check(self, i, index):
        # Deep check based on index and previous CRC32 values to verify the assumption
        string = ""
        tc = 0x00
        hashcode = self.crc32(i)
        tc = hashcode & 0xff ^ index[2]
        if not (tc <= 57 and tc >= 48):
            return [0]
        string += str(tc - 48)
        hashcode = self.crctable[index[2]] ^ (hashcode >> 8)
        tc = hashcode & 0xff ^ index[1]
        if not (tc <= 57 and tc >= 48):
            return [0]
        string += str(tc - 48)
        hashcode = self.crctable[index[1]] ^ (hashcode >> 8)
        tc = hashcode & 0xff ^ index[0]
        if not (tc <= 57 and tc >= 48):
            return [0]
        string += str(tc - 48)
        hashcode = self.crctable[index[0]] ^ (hashcode >> 8)
        return [1, string]

    def main(self, string):
        # Main function to compute and validate CRC32 for the given string
        index = [0 for x in range(4)]
        i = 0
        ht = int(f"0x{string}", 16) ^ 0xffffffff
        for i in range(3, -1, -1):
            index[3-i] = self.get_crc_index(ht >> (i*8))
            snum = self.crctable[index[3-i]]
            ht ^= snum >> ((3-i)*8)
        for i in range(100000000):
            lastindex = self.crc32_last_index(i)
            if lastindex == index[3]:
                deepCheckData = self.deep_check(i, index)
                if deepCheckData[0]:
                    break
        if i == 100000000:
            return -1
        return f"{i}{deepCheckData[1]}"
class Tools:
    @staticmethod
    # Get save path and format
    def get_save():
        return os.path.join(os.path.join(os.path.expanduser("~"), "Desktop"),
                            "Bilibili_Video_Analysis_{}.xlsx".format(datetime.now().strftime('%Y-%m-%d')))

    @staticmethod
    # Format timestamp
    def format_timestamp(timestamp):
        dt_object = datetime.fromtimestamp(timestamp)
        formatted_time = dt_object.strftime("%Y-%m-%d %H:%M:%S")
        return formatted_time

    @staticmethod
    # Calculate sentiment score
    def calculate_sentiment_score(text):
        s = SnowNLP(text)
        sentiment_score = s.sentiments
        return sentiment_score

    @staticmethod
    # Generate a word cloud
    def get_word_cloud(sheet_name: str, workbook: Workbook):
        sheet = workbook[sheet_name]

        # Read frequency data
        words = []
        frequencies = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            words.append(row[0])
            frequencies.append(row[1])

        system = platform.system()

        if system == 'Darwin':  # macOS
            font_path = '/System/Library/Fonts/STHeiti Light.ttc'
        elif system == 'Windows':
            font_path = 'C:/Windows/Fonts/simhei.ttf'
        else:  # Other OS
            font_path = 'simhei.ttf'

        wordcloud = WordCloud(background_color='white', max_words=100, font_path=font_path)
        word_frequency = dict(zip(words, frequencies))
        wordcloud.generate_from_frequencies(word_frequency)

        plt.imshow(wordcloud, interpolation='bilinear')
        plt.axis('off')
        plt.show()

    @staticmethod
    # Generate horizontal statistics chart
    def get_word_chart(sheet_name: str, workbook):
        sheet = workbook[sheet_name]

        words = []
        frequencies = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            words.append(row[0])
            frequencies.append(row[1])

        system = platform.system()

        if system == 'Darwin':  
            font_path = '/System/Library/Fonts/STHeiti Light.ttc'
        elif system == 'Windows':
            font_path = 'C:/Windows/Fonts/simhei.ttf'
        else:  
            font_path = 'simhei.ttf'

        custom_font = fm.FontProperties(fname=font_path)

        fig, ax = plt.subplots()
        ax.barh(words, frequencies)
        ax.set_xlabel("Frequency", fontproperties=custom_font)
        ax.set_ylabel("Words", fontproperties=custom_font)

        plt.yticks(fontproperties=custom_font)

        plt.show()

    @staticmethod
    def get_user_info_by_card(user_card_json):
        info = {
            'name': "N/A", 'birthday': "N/A", 'regtime': "N/A",
            'fans': "N/A", 'friend': "N/A"
        }

        try:
            info['name'] = user_card_json['card']['name']
            info['birthday'] = user_card_json['card']['birthday']
            info['regtime'] = Tools.format_timestamp(int(user_card_json['card']['regtime']))
            info['fans'] = user_card_json['card']['fans']
            info['friend'] = user_card_json['card']['friend']
        except KeyError:
            pass

        return tuple(info.values())

class BilibiliExcel:
    @staticmethod
    # Write video basic information
    def write_base_info(workbook, bv_json):
        sheet = workbook.create_sheet(title="Video Info")
        headers = ["Video Title", "Author", "Publish Time", "Views", "Favorites", "Shares", "Total Bullet Comments",
                   "Comments Count", "Video Description", "Category", "Video Link", "Thumbnail Link"]
        sheet.append(headers)

        data = [bv_json["data"]["title"],
                bv_json["data"]["owner"]["name"],
                Tools.format_timestamp(bv_json["data"]["pubdate"]),
                bv_json["data"]["stat"]["view"],
                bv_json["data"]["stat"]["favorite"],
                bv_json["data"]["stat"]["share"],
                bv_json["data"]["stat"]["danmaku"],
                bv_json["data"]["stat"]["reply"],
                bv_json["data"]["desc"],
                bv_json["data"]["tname"],
                video_url,
                bv_json["data"]["pic"]]

        sheet.append(data)

    @staticmethod
    def save_workbook(workbook):
        workbook.save(Tools.get_save())

class PrintInfo:
    # Print basic information
    @staticmethod
    def base_message():
        if 'Windows' == platform.system():
            os.system('cls')
        else:
            os.system('clear')

        text = '''
        ************************************

        Bilibili Video Analysis v2023.6.26
        Author: Github.com/hoochanlon
        Project URL: https://github.com/hoochanlon/scripts

        Features:
        1. Analyze and visualize Bilibili video data.

        Disclaimer: For research and learning purposes only.

        ************************************
        '''
        print(text.center(50, ' '))

if __name__ == '__main__':
    PrintInfo.base_message()

    while True:
        video_url = input("Paste the Bilibili video link: ")
        if re.match(r'.*BV\w+', video_url):
            break
        else:
            print("Invalid link format. Please re-enter.")

    bv_json = BilibiliAPI.get_bv_json(video_url)
    workbook = Workbook()
    workbook.remove(workbook.active)
    BilibiliExcel.write_base_info(workbook, bv_json)
    BilibiliExcel.save_workbook(workbook)
