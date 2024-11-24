import os
import requests
import json
from openpyxl import Workbook
from datetime import datetime

# Your existing get_citycode and get_job_list functions should already be defined in your script.
# Make sure they are present in the script or imported from another module.

# For example, let's assume this is part of your script:
def get_citycode(city_name):
    response = requests.get('https://api.hh.ru/areas')
    city_data = response.json()
    
    # Ищем в каждом регионе города
    for region in city_data:
        # Если город найден в регионе, возвращаем его ID
        for area in region.get('areas', []):
            if area['name'] == city_name:
                return area['id']
    return None

def get_job_list(keywords, city_name):
    city_code = get_citycode(city_name)
    if city_code is None:
        print(f"Город '{city_name}' не найден.")
        return []

    url = "https://api.hh.ru/vacancies"
    params = {
        'text': keywords,  # Ключевые слова для поиска вакансий
        'area': city_code,  # ID города
        'page': 0,  # Стартовая страница
        'per_page': 10,  # Количество вакансий на странице
    }

    response = requests.get(url, params=params)
    data = response.json()

    return data['items']

# Your DataToExcel class remains the same.

class DataToExcel:
    @staticmethod
    def save_to_excel(workbook, hhru_json):
        # Load the hh.ru JSON data
        with open(hhru_json, 'r', encoding='utf-8') as f:
            hhru_json = json.load(f)

        worksheet = workbook.create_sheet(title='Job Listings')

        # Define the column titles
        titles = [
            'Job Title', 'Skills', 'Experience', 'Salary From', 'Salary To', 'Currency', 
            'Company', 'Employment Type', 'Experience Required', 'City', 'Job URL'
        ]

        # Write the titles in the first row
        for col, title in enumerate(titles, start=1):
            worksheet.cell(row=1, column=col, value=title)

        # Extract job listings from the JSON data
        job_data_list = hhru_json.get('items', [])
        row = 2  # Start from the second row for data

        for job in job_data_list:
            worksheet.cell(row=row, column=1, value=job.get('name', 'N/A'))
            worksheet.cell(row=row, column=2, value=', '.join(job.get('key_skills', [])))
            worksheet.cell(row=row, column=3, value=job.get('experience', {}).get('name', 'N/A'))
            
            # Handle salary data
            salary = job.get('salary', {})
            salary_from = salary.get('from', 'N/A') if salary else 'N/A'
            salary_to = salary.get('to', 'N/A') if salary else 'N/A'
            salary_currency = salary.get('currency', 'N/A') if salary else 'N/A'
            
            worksheet.cell(row=row, column=4, value=salary_from)
            worksheet.cell(row=row, column=5, value=salary_to)
            worksheet.cell(row=row, column=6, value=salary_currency)
            
            worksheet.cell(row=row, column=7, value=job.get('employer', {}).get('name', 'N/A'))
            worksheet.cell(row=row, column=8, value=job.get('employment', {}).get('name', 'N/A'))
            worksheet.cell(row=row, column=9, value=job.get('experience', {}).get('name', 'N/A'))
            worksheet.cell(row=row, column=10, value=job.get('area', {}).get('name', 'N/A'))
            worksheet.cell(row=row, column=11, value=job.get('alternate_url', 'N/A'))
            
            row += 1  # Move to the next row

        # Save the workbook to a file
        save_xlsx = os.path.join(
            os.path.expanduser("~"), "Desktop", f"hhru_jobs_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )
        workbook.save(save_xlsx)
        print(f"Job listings have been saved to: {save_xlsx}")


# Example function that saves the job data and calls the Excel export function
def save_hhru_data_to_excel():
    # Ensure get_job_list is available here.
    job_listings = get_job_list("разработчик, кибербезопасность, IT", "Москва")
    
    # Save the job listings to a JSON file
    hhru_json_path = os.path.join(os.path.expanduser("~"), "Desktop", f"hhru_jobs_{datetime.now().strftime('%Y-%m-%d')}.json")
    with open(hhru_json_path, 'w', encoding='utf-8') as f:
        json.dump({"items": job_listings}, f, ensure_ascii=False, indent=4)
    
    # Create a new workbook and save the data to Excel
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove the default sheet
    DataToExcel.save_to_excel(workbook, hhru_json_path)


# Call the function to save the job listings to Excel
save_hhru_data_to_excel()
