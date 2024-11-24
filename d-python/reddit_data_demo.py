import praw
from openpyxl import Workbook
import os
from datetime import datetime

class Tools:
    @staticmethod
    def init_reddit():
        """Initialize the Reddit client using PRAW."""
        reddit = praw.Reddit(
            client_id='YOUR_CLIENT_ID',  # Replace with your Reddit app's client ID
            client_secret='YOUR_CLIENT_SECRET',  # Replace with your Reddit app's client secret
            user_agent='Reddit Data Export Script'  # A custom user agent for the script
        )
        return reddit

    @staticmethod
    def fetch_post_data(reddit, subreddit_name, post_id):
        """Fetch Reddit post and its comments."""
        subreddit = reddit.subreddit(subreddit_name)
        post = reddit.submission(id=post_id)
        post.comments.replace_more(limit=None)  # Fetch all comments, removing pagination
        return post

    @staticmethod
    def format_timestamp(timestamp):
        """Format a Unix timestamp into a human-readable string."""
        dt_object = datetime.fromtimestamp(int(timestamp))
        return dt_object.strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def get_save():
        """Generate the file path to save the Excel file."""
        return os.path.join(
            os.path.join(os.path.expanduser("~"), "Desktop"),
            "Reddit_Post_{}.xlsx".format(datetime.now().strftime('%Y-%m-%d'))
        )

class RedditExcel:
    @staticmethod
    def init_info(workbook, post):
        """Create and populate an Excel sheet with post and comment details."""
        worksheet = workbook.create_sheet(title='Reddit Post Details')

        # Write headers
        headers = [
            'Post Title', 'Post Author', 'Post Score', 'Post Comments Count',
            'Post URL', 'Comment Author', 'Comment Body', 'Comment Score', 'Comment Time'
        ]
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)

        # Write post details
        row = 2
        worksheet.cell(row=row, column=1, value=post.title)
        worksheet.cell(row=row, column=2, value=post.author.name if post.author else 'N/A')
        worksheet.cell(row=row, column=3, value=post.score)
        worksheet.cell(row=row, column=4, value=post.num_comments)
        worksheet.cell(row=row, column=5, value=post.url)

        # Write comments
        for comment in post.comments.list():
            if isinstance(comment, praw.models.Comment):
                row += 1
                worksheet.cell(row=row, column=6, value=comment.author.name if comment.author else 'N/A')
                worksheet.cell(row=row, column=7, value=comment.body)
                worksheet.cell(row=row, column=8, value=comment.score)
                worksheet.cell(row=row, column=9, value=Tools.format_timestamp(comment.created_utc))

if __name__ == '__main__':
    # Initialize Reddit API
    reddit = Tools.init_reddit()

    # Get subreddit name and post ID from the user
    while True:
        subreddit_name = input("Enter subreddit name (e.g., 'python'): ")
        post_id = input("Enter Reddit post ID (found in the URL): ")
        try:
            post = Tools.fetch_post_data(reddit, subreddit_name, post_id)
            break
        except Exception as e:
            print(f"Error fetching post: {e}. Please try again.")

    # Create a workbook and remove the default sheet
    workbook = Workbook()
    workbook.remove(workbook.active)

    # Initialize Excel writing
    RedditExcel.init_info(workbook, post)

    # Save the workbook
    file_path = Tools.get_save()
    workbook.save(file_path)
    print(f"Data saved to {file_path}")
